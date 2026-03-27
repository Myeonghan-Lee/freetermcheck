import streamlit as st
import pandas as pd
import openpyxl
import re
from io import BytesIO

# --- 새로 추가된 openpyxl 스타일 및 서식 모듈 ---
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText


def extract_numbers_from_bracket(text):
    """문자열에서 괄호 안의 숫자를 찾아 합산합니다."""
    if not text:
        return 0
    numbers = re.findall(r'\((\d+)\)', str(text))
    return sum(int(n) for n in numbers)

def evaluate_formula_string(text):
    """문자열에서 숫자와 기호만 추출하여 계산합니다."""
    if not text:
        return 0
    try:
        clean_expr = re.sub(r'[^\d\.\*\+\-\/]', '', str(text))
        if clean_expr:
            return eval(clean_expr)
    except:
        return 0
    return 0

def process_file(file):
    results = []
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        filename = file.name
        
        # --- 1. 학교운영 현황 검토 ---
        ws1 = wb.get_sheet_by_name("1.학교운영 현황") if "1.학교운영 현황" in wb.sheetnames else None
        if ws1:
            theme_hours = ws1['D8'].value or 0
            career_hours = ws1['D9'].value or 0
            class_count = ws1['C11'].value or 0
            
            e8_text = ws1['E8'].value
            e9_text = ws1['E9'].value
            
            if extract_numbers_from_bracket(e8_text) != theme_hours:
                results.append(f"[1.학교운영 현황] D8(주제선택 시수: {theme_hours})와 E8 병합셀의 과목 시수 합이 불일치합니다.")
            if extract_numbers_from_bracket(e9_text) != career_hours:
                results.append(f"[1.학교운영 현황] D9(진로탐색 시수: {career_hours})와 E9 병합셀의 과목 시수 합이 불일치합니다.")
        else:
            results.append("[1.학교운영 현황] 시트를 찾을 수 없습니다.")

        # --- 2. 자유학기 활동 검토 ---
        ws2 = wb.get_sheet_by_name("2. 자유학기 활동") if "2. 자유학기 활동" in wb.sheetnames else None
        has_outsourced = False
        
        if ws2:
            current_section = None
            total_theme_hours = 0
            total_career_hours = 0
            
            for row in range(5, ws2.max_row + 1):
                cell_val = str(ws2.cell(row=row, column=1).value or ws2.cell(row=row, column=2).value or "")
                
                if '주제선택 활동' in cell_val:
                    current_section = 'theme'
                    continue
                elif '진로 탐색 활동' in cell_val:
                    current_section = 'career'
                    continue
                    
                g_val = ws2.cell(row=row, column=7).value
                e_val = ws2.cell(row=row, column=5).value
                
                if isinstance(g_val, (int, float)):
                    if current_section == 'theme':
                        total_theme_hours += g_val
                    elif current_section == 'career':
                        total_career_hours += g_val
                        
                if e_val and '개인위탁' in str(e_val):
                    has_outsourced = True

            required_theme = theme_hours * class_count
            required_career = career_hours * class_count
            
            if total_theme_hours < required_theme:
                results.append(f"[2. 자유학기 활동] 주제선택 총 시수({total_theme_hours})가 기준({required_theme})보다 부족합니다.")
            if total_career_hours < required_career:
                results.append(f"[2. 자유학기 활동] 진로탐색 총 시수({total_career_hours})가 기준({required_career})보다 부족합니다.")
        else:
            results.append("[2. 자유학기 활동] 시트를 찾을 수 없습니다.")

        # --- 3. 예산 계획서 검토 ---
        ws3 = wb.get_sheet_by_name("3. 예산 계획서") if "3. 예산 계획서" in wb.sheetnames else None
        if ws3:
            total_budget = ws3['E3'].value or 0
            
            for row in range(6, 31):
                b_val = ws3.cell(row=row, column=2).value
                c_val = ws3.cell(row=row, column=3).value
                d_val = ws3.cell(row=row, column=4).value
                
                if row == 17:
                    if str(b_val).strip() != '프로그램 개인위탁 운영비':
                        results.append("[3. 예산 계획서] B17 셀의 내용이 '프로그램 개인위탁 운영비'가 아닙니다.")
                    if has_outsourced:
                        if not c_val or not d_val:
                            results.append("[3. 예산 계획서] 개인위탁 교사가 있으나 17행의 산출근거 또는 소요예산이 누락되었습니다.")
                    continue
                
                if b_val:
                    if not c_val:
                        results.append(f"[3. 예산 계획서] {row}행: 내용이 있으나 산출근거가 없습니다.")
                    else:
                        calculated_budget = evaluate_formula_string(c_val)
                        if calculated_budget > 0 and d_val:
                            if abs(calculated_budget - d_val) > 10: 
                                results.append(f"[3. 예산 계획서] {row}행: 산출근거 계산값과 소요예산이 일치하지 않습니다. (입력값: {d_val})")
            
            biz_expense = ws3['D31'].value or 0
            if biz_expense > (total_budget * 0.03):
                results.append(f"[3. 예산 계획서] D31 업무추진비({biz_expense})가 총 예산의 3%를 초과합니다.")
                
        else:
            results.append("[3. 예산 계획서] 시트를 찾을 수 없습니다.")

        if not results:
            results.append("특이사항 없음 (모든 검토 항목을 통과했습니다.)")

    except Exception as e:
        results.append(f"파일을 읽는 중 오류가 발생했습니다: {e}")
        
    return filename, results

# --- HTML 스타일 적용 함수 ---
def format_issue_for_html(issue):
    """시트 이름별로 다른 글자 색상을 적용하고 HTML 태그로 감쌉니다."""
    if "[1.학교운영 현황]" in issue:
        issue = issue.replace("[1.학교운영 현황]", "<strong style='color:#0052cc;'>[1.학교운영 현황]</strong>")
    elif "[2. 자유학기 활동]" in issue:
        issue = issue.replace("[2. 자유학기 활동]", "<strong style='color:#00875a;'>[2. 자유학기 활동]</strong>")
    elif "[3. 예산 계획서]" in issue:
        issue = issue.replace("[3. 예산 계획서]", "<strong style='color:#de350b;'>[3. 예산 계획서]</strong>")
    return f"• {issue}"

# --- Streamlit UI ---
st.set_page_config(page_title="자유학기 운영계획서 검토기", layout="wide")

# 테이블 전체 넓이 및 테두리 스타일을 위한 CSS 주입
st.markdown("""
<style>
table { width: 100%; border-collapse: collapse; }
th, td { text-align: left; padding: 12px; border: 1px solid #ddd; line-height: 1.6; }
</style>
""", unsafe_allow_html=True)

st.title("📄 자유학기 운영계획서 자동 검토 웹앱")
st.write("여러 개의 엑셀 파일(.xlsx)을 업로드하면 운영계획서 작성 지침에 맞는지 자동으로 검토합니다.")

uploaded_files = st.file_uploader("검토할 엑셀 파일들을 업로드하세요", type=['xlsx'], accept_multiple_files=True)

if uploaded_files:
    if st.button("검토 시작"):
        report_data = []
        
        with st.spinner("파일을 검토하는 중입니다..."):
            for file in uploaded_files:
                filename, issues = process_file(file)
                
                # 1. 화면 출력용 (HTML, 줄바꿈 <br> 적용, 시트별 색상 적용)
                html_issues = "<br>".join([format_issue_for_html(issue) for issue in issues])
                
                # 2. 엑셀 다운로드용 (순수 텍스트, 줄바꿈 \n 적용)
                excel_issues = "\n".join([f"• {issue}" for issue in issues])
                
                # 3. 이상 유무 판별 (성공 여부 확인)
                is_success = "특이사항 없음" in "".join(issues)
                
                report_data.append({
                    "파일명": filename, 
                    "검토 결과 (화면용)": html_issues,
                    "검토 결과 (Excel용)": excel_issues,
                    "is_success": is_success
                })
        
        # 이상이 없는 파일이 맨 위에 오도록 내림차순 정렬
        report_data.sort(key=lambda x: x["is_success"], reverse=True)
        
        # DataFrame 생성
        df_results = pd.DataFrame(report_data)
        
        # 화면에 표시할 칼럼만 추출
        display_df = df_results[['파일명', '검토 결과 (화면용)']].copy()
        display_df.rename(columns={'검토 결과 (화면용)': '검토 결과'}, inplace=True)
        
        # 이상 없는 파일명에 초록색 배경을 넣는 조건부 서식
        def highlight_success(row):
            if "특이사항 없음" in row['검토 결과']:
                return ['background-color: #e6ffe6; color: #006600; font-weight: bold', '']
            else:
                return [''] * len(row)
        
        # 스타일 적용 및 HTML 테이블로 변환
        styled_df = display_df.style.apply(highlight_success, axis=1)
        html_table = styled_df.hide(axis="index").to_html(escape=False)
        
        st.subheader("📊 검토 결과")
        # 변환된 HTML 테이블을 화면에 출력
        st.markdown(html_table, unsafe_allow_html=True)
        
        st.write("---")
        
        # --- 엑셀 파일 다운로드 기능 (스타일 적용) ---
        download_df = df_results[['파일명', '검토 결과 (Excel용)']].copy()
        download_df.rename(columns={'검토 결과 (Excel용)': '검토 결과'}, inplace=True)
        
        # 메모리 버퍼 생성 후 엑셀 데이터 쓰기
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            download_df.to_excel(writer, index=False, sheet_name='검토결과')
            
            worksheet = writer.sheets['검토결과']
            
            # 1. 헤더 스타일 적용 (굵게, 옅은 회색 배경)
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 2. Rich Text 폰트 설정 (ARGB 포맷)
            font_1 = InlineFont(color='FF0052CC', b=True) # [1.학교운영 현황] - 파란색
            font_2 = InlineFont(color='FF00875A', b=True) # [2. 자유학기 활동] - 초록색
            font_3 = InlineFont(color='FFDE350B', b=True) # [3. 예산 계획서] - 빨간색

            # 3. 데이터 셀 순회 및 스타일 적용
            for row in worksheet.iter_rows(min_row=2, max_col=2):
                filename_cell = row[0]
                result_cell = row[1]
                
                text_val = result_cell.value
                if not text_val: 
                    continue
                
                # 특이사항이 없는 경우 (화면과 동일하게 행 전체 초록색 배경 처리)
                if "특이사항 없음" in text_val:
                    success_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                    success_font = Font(color="006600", bold=True)
                    filename_cell.fill = success_fill
                    filename_cell.font = success_font
                    result_cell.fill = success_fill
                    result_cell.font = success_font
                
                # 문제가 있는 경우 (항목별 태그 색상 변경)
                else:
                    lines = text_val.split('\n')
                    rich_elements = []
                    
                    for i, line in enumerate(lines):
                        if "[1.학교운영 현황]" in line:
                            idx = line.find("[1.학교운영 현황]") + len("[1.학교운영 현황]")
                            rich_elements.append(TextBlock(font=font_1, text=line[:idx]))
                            if line[idx:]: rich_elements.append(line[idx:])
                        
                        elif "[2. 자유학기 활동]" in line:
                            idx = line.find("[2. 자유학기 활동]") + len("[2. 자유학기 활동]")
                            rich_elements.append(TextBlock(font=font_2, text=line[:idx]))
                            if line[idx:]: rich_elements.append(line[idx:])
                        
                        elif "[3. 예산 계획서]" in line:
                            idx = line.find("[3. 예산 계획서]") + len("[3. 예산 계획서]")
                            rich_elements.append(TextBlock(font=font_3, text=line[:idx]))
                            if line[idx:]: rich_elements.append(line[idx:])
                        
                        else:
                            if line: rich_elements.append(line)
                        
                        # 줄바꿈 유지
                        if i < len(lines) - 1:
                            rich_elements.append("\n")
                    
                    if rich_elements:
                        result_cell.value = CellRichText(rich_elements)
                
                # 셀 내용 자동 줄바꿈 및 위쪽 정렬 적용
                filename_cell.alignment = Alignment(wrapText=True, vertical='top')
                result_cell.alignment = Alignment(wrapText=True, vertical='top')
            
            # 4. 열 너비 조절 (가독성 향상)
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 100
            
        # 다운로드 버튼
        st.download_button(
            label="📥 검토 결과 다운로드 (Excel)",
            data=excel_buffer.getvalue(),
            file_name="운영계획서_검토결과.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
